import yfinance as yf
import sqlite3
from datetime import datetime
import logging
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dataclasses import dataclass
from typing import Dict, List

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s: %(message)s')

@dataclass
class Trade:
    symbol: str
    action: str  # 'BUY', 'SELL', 'SHORT', 'COVER'
    quantity: float
    price: float
    timestamp: datetime
    commission: float
    trade_id: str
    currency: str  # 'GBX' or 'USD'

@dataclass
class Position:
    symbol: str
    quantity: float
    avg_cost: float
    current_price: float
    market_value: float
    unrealized_pnl: float
    realized_pnl: float
    position_type: str  # LONG or SHORT
    currency: str       # 'GBX' or 'USD'

class PaperTradingSystem:
    def __init__(self, initial_cash_gbp: float = 6000.0, commission_rate: float = 0.005, reset_on_start=False):
        self.commission_rate = commission_rate
        self.min_commission_gbp = 1.00
        self.min_commission_usd = 1.00
        self.margin_requirement = 0.5  # For short selling margin calculation
        
        self.db_path = 'paper_trading.db'
        self.excel_path = 'paper_trading_portfolio.xlsx'
        self.cash_gbp = initial_cash_gbp    # Single cash in GBP only
        
        if reset_on_start:
            if os.path.exists(self.db_path): os.remove(self.db_path)
            if os.path.exists(self.excel_path): os.remove(self.excel_path)
        
        self.init_database()
        
        self.positions: Dict[str, Position] = {}
        self.trade_history: List[Trade] = []
        self.market_data_cache = {}
        self.last_update = None
        
        self.load_portfolio_from_db()
        self.init_excel_file()
    
    def init_database(self):
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute('''
                CREATE TABLE IF NOT EXISTS trades (
                    trade_id TEXT PRIMARY KEY, symbol TEXT, action TEXT, quantity REAL, price REAL,
                    commission REAL, currency TEXT, timestamp TEXT
                )''')
            cur.execute('''
                CREATE TABLE IF NOT EXISTS portfolio (
                    symbol TEXT PRIMARY KEY, quantity REAL, avg_cost REAL, realized_pnl REAL,
                    position_type TEXT, currency TEXT
                )''')
            cur.execute('''
                CREATE TABLE IF NOT EXISTS account (
                    cash_gbp REAL, last_updated TEXT
                )''')
            cur.execute('SELECT COUNT(*) FROM account')
            if cur.fetchone()[0] == 0:
                cur.execute('INSERT INTO account (cash_gbp, last_updated) VALUES (?, ?)',
                            (self.cash_gbp, datetime.now().isoformat()))
    
    def init_excel_file(self):
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws1 = wb.active
            ws1.title = 'Trades'
            ws1.append(['Trade ID', 'Symbol', 'Action', 'Quantity', 'Price', 'Commission', 'Value', 'Currency', 'Timestamp'])
            ws2 = wb.create_sheet('Portfolio')
            ws2.append(['Symbol', 'Quantity', 'Avg Cost', 'Current Price', 'Market Value',
                        'Unrealized PnL', 'Realized PnL', 'Position Type', 'Currency'])
            ws3 = wb.create_sheet('Summary')
            ws3.append(['Metric', 'Value', 'Currency', 'Last Updated'])
            self._style_header(ws1)
            self._style_header(ws2)
            self._style_header(ws3)
            wb.save(self.excel_path)

    def _style_header(self, worksheet):
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    def get_fx_rate_usd_to_gbp(self) -> float:
        """Fetch USD->GBP FX rate by inverting GBPUSD=X"""
        try:
            ticker = yf.Ticker("GBPUSD=X")
            data = ticker.history(period='1d')
            if data.empty:
                logging.warning("GBPUSD FX rate not available, using fallback 0.78")
                return 0.78
            gbp_usd = data['Close'].iloc[-1]
            usd_gbp = 1 / gbp_usd
            return usd_gbp
        except Exception as e:
            logging.error(f"Error fetching FX rate USD->GBP: {e}")
            return 0.78

    def get_current_price(self, symbol: str):
        """
        Fetch current price and infer currency:
          - UK stocks with .L suffix or known UK tickers assumed GBX (prices in pence)
          - US stocks assumed USD
        """
        symbol_upper = symbol.upper()
        is_uk_stock = symbol_upper.endswith('.L')
        try:
            ticker = yf.Ticker(symbol if is_uk_stock else symbol_upper)
            hist = ticker.history(period="1d", interval="1m")
            if hist.empty:
                hist = ticker.history(period="5d")
            if hist.empty:
                logging.warning(f"No price data for {symbol}")
                return 0.0, 'UNKNOWN'
            price = hist['Close'].iloc[-1]

            info = ticker.info or {}
            currency = info.get('currency', None)

            if is_uk_stock:
                # Treat UK stocks as GBX (pence sterling)
                currency = 'GBX'
            else:
                if currency is None:
                    currency = 'USD'  # default fallback
            return float(price), currency
        except Exception as e:
            logging.error(f"Failed fetching price for {symbol}: {e}")
            return 0.0, 'UNKNOWN'

    def update_market_data(self):
        for symbol in self.positions.keys():
            price, currency = self.get_current_price(symbol)
            if price > 0:
                pos = self.positions[symbol]
                pos.current_price = price
                pos.currency = currency  # update currency in case it differs
                
                if pos.position_type == 'LONG':
                    pos.market_value = pos.quantity * pos.current_price
                    pos.unrealized_pnl = (pos.current_price - pos.avg_cost) * pos.quantity
                elif pos.position_type == 'SHORT':
                    pos.market_value = abs(pos.quantity) * pos.current_price
                    pos.unrealized_pnl = (pos.avg_cost - pos.current_price) * abs(pos.quantity)
        self.last_update = datetime.now()

    def calculate_commission(self, trade_value: float, currency: str) -> float:
        commission = trade_value * self.commission_rate
        if currency == 'GBX':
            # commission is GBX, min commission £1 == 100 GBX
            min_commission_gbx = self.min_commission_gbp * 100
            return max(commission, min_commission_gbx)
        elif currency == 'USD':
            return max(commission, self.min_commission_usd)
        else:
            # Use GBP min commission fallback
            return max(commission, self.min_commission_gbp)

    def place_order(self, symbol: str, action: str, quantity: float) -> bool:
        action = action.upper()
        if quantity <= 0 or action not in ['BUY', 'SELL', 'SHORT', 'COVER']:
            logging.error("Invalid action or quantity")
            return False
        qty = float(quantity)
        price, currency = self.get_current_price(symbol)
        if price <= 0:
            logging.error(f"Invalid price for symbol {symbol}")
            return False
        
        trade_value = qty * price
        commission = self.calculate_commission(trade_value, currency)
        usd_to_gbp = self.get_fx_rate_usd_to_gbp()

        # Cash impact for BUY and COVER = cash outflow, needs funds
        # Cash impact for SELL and SHORT = cash inflow (proceeds minus commission)
        if action in ['BUY', 'COVER']:
            # Calculate cash needed in GBP for this trade
            if currency == 'GBX':
                # Convert GBX to GBP by dividing by 100
                required_cash = (trade_value + commission) / 100
            elif currency == 'USD':
                required_cash = (trade_value + commission) * usd_to_gbp
            else:
                logging.error(f"Unsupported currency: {currency}")
                return False
            
            if required_cash > self.cash_gbp:
                logging.error(f"Insufficient cash. Need £{required_cash:.2f}, available £{self.cash_gbp:.2f}")
                return False
            self.cash_gbp -= required_cash
        elif action in ['SELL', 'SHORT']:
            # Cash inflow in GBP
            if currency == 'GBX':
                proceeds_gbp = (trade_value - commission) / 100
            elif currency == 'USD':
                proceeds_gbp = (trade_value - commission) * usd_to_gbp
            else:
                logging.error(f"Unsupported currency: {currency}")
                return False
            
            self.cash_gbp += proceeds_gbp

        pos = self.positions.get(symbol)
        
        if action == 'BUY':
            if pos and pos.position_type == 'LONG' and pos.currency == currency:
                total_qty = pos.quantity + qty
                total_cost = pos.avg_cost * pos.quantity + price * qty
                pos.avg_cost = total_cost / total_qty
                pos.quantity = total_qty
            elif pos and pos.position_type == 'SHORT':
                # Cover shorts on buy
                cover_qty = min(qty, abs(pos.quantity))
                realized = (pos.avg_cost - price) * cover_qty - commission / (100 if currency == 'GBX' else 1)
                pos.realized_pnl += realized * (1 if currency == 'GBX' else usd_to_gbp)
                pos.quantity += cover_qty
                remaining = qty - cover_qty
                if abs(pos.quantity) < 1e-6:
                    del self.positions[symbol]
                if remaining > 1e-6:
                    # Open new long position with remaining qty
                    self.positions[symbol] = Position(
                        symbol=symbol,
                        quantity=remaining,
                        avg_cost=price,
                        current_price=price,
                        market_value=remaining * price,
                        unrealized_pnl=0.0,
                        realized_pnl=0.0,
                        position_type='LONG',
                        currency=currency
                    )
            else:
                self.positions[symbol] = Position(
                    symbol=symbol,
                    quantity=qty,
                    avg_cost=price,
                    current_price=price,
                    market_value=trade_value,
                    unrealized_pnl=0.0,
                    realized_pnl=0.0,
                    position_type='LONG',
                    currency=currency
                )
        elif action == 'SELL':
            if not pos or pos.position_type != 'LONG' or pos.quantity < qty:
                logging.error("Trying to sell more than held or no long position")
                return False
            cost_basis = pos.avg_cost * qty
            realized = (price * qty - cost_basis) - commission / (100 if currency == 'GBX' else 1)
            pos.realized_pnl += realized * (1 if currency == 'GBX' else usd_to_gbp)
            pos.quantity -= qty
            if pos.quantity < 1e-6:
                del self.positions[symbol]
        elif action == 'SHORT':
            if currency != 'GBX':
                logging.error("Short sales only supported for UK stocks in GBX currency")
                return False
            margin_required = trade_value * self.margin_requirement
            margin_required_gbp = margin_required / 100
            if margin_required_gbp > self.cash_gbp:
                logging.error(f"Insufficient cash for margin £{margin_required_gbp:.2f}")
                return False
            # Margin accounted elsewhere or as simple cash deduction here
            pos_qty = abs(pos.quantity) if pos and pos.position_type == 'SHORT' else 0.0
            total_qty = pos_qty + qty
            if pos and pos.position_type == 'SHORT':
                total_cost = pos.avg_cost * pos_qty + price * qty
                avg_cost = total_cost / total_qty
                pos.avg_cost = avg_cost
                pos.quantity = -total_qty
            else:
                if pos and pos.position_type == 'LONG':
                    # You could allow closing long position before shorting; add logic as needed
                    logging.error("Cannot short when holding long position")
                    return False
                self.positions[symbol] = Position(
                    symbol=symbol,
                    quantity=-qty,
                    avg_cost=price,
                    current_price=price,
                    market_value=trade_value,
                    unrealized_pnl=0.0,
                    realized_pnl=0.0,
                    position_type='SHORT',
                    currency=currency
                )
        elif action == 'COVER':
            if not pos or pos.position_type != 'SHORT' or abs(pos.quantity) < qty:
                logging.error("Trying to cover more than short held or no short position")
                return False
            realized = (pos.avg_cost - price) * qty - commission / (100 if currency == 'GBX' else 1)
            pos.realized_pnl += realized * (1 if currency == 'GBX' else usd_to_gbp)
            pos.quantity += qty
            if abs(pos.quantity) < 1e-6:
                del self.positions[symbol]

        self._record_trade(symbol, action, qty, price, commission, currency)
        self.save_to_db()
        return True

    def _record_trade(self, symbol, action, qty, price, commission, currency):
        trade_id = f"{symbol}_{action}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
        trade = Trade(symbol, action, qty, price, datetime.now(), commission, trade_id, currency)
        self.trade_history.append(trade)

    def save_to_db(self):
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute('DELETE FROM account')
            cur.execute('INSERT INTO account (cash_gbp, last_updated) VALUES (?, ?)',
                        (self.cash_gbp, datetime.now().isoformat()))
            cur.execute('DELETE FROM portfolio')
            for p in self.positions.values():
                if abs(p.quantity) > 1e-6:
                    cur.execute('INSERT INTO portfolio (symbol, quantity, avg_cost, realized_pnl, position_type, currency) VALUES (?, ?, ?, ?, ?, ?)',
                                (p.symbol, p.quantity, p.avg_cost, p.realized_pnl, p.position_type, p.currency))
            cur.execute('DELETE FROM trades')
            for t in self.trade_history:
                cur.execute('INSERT INTO trades (trade_id, symbol, action, quantity, price, commission, currency, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                            (t.trade_id, t.symbol, t.action, t.quantity, t.price, t.commission, t.currency, t.timestamp.isoformat()))
            conn.commit()
            self.export_to_excel()

    def load_portfolio_from_db(self):
        if not os.path.exists(self.db_path):
            return
        with sqlite3.connect(self.db_path) as conn:
            cur = conn.cursor()
            cur.execute('SELECT cash_gbp FROM account ORDER BY last_updated DESC LIMIT 1')
            r = cur.fetchone()
            self.cash_gbp = r[0] if r else 6000.0
            cur.execute('SELECT symbol, quantity, avg_cost, realized_pnl, position_type, currency FROM portfolio')
            self.positions.clear()
            for row in cur.fetchall():
                sym, qty, avg_cost, r_pnl, p_type, curc = row
                self.positions[sym] = Position(sym, qty, avg_cost, 0.0, 0.0, 0.0, r_pnl, p_type, curc)
            cur.execute('SELECT trade_id, symbol, action, quantity, price, commission, currency, timestamp FROM trades ORDER BY timestamp')
            self.trade_history.clear()
            for t in cur.fetchall():
                trade_id, symbol, action, qty, price, com, curc, ts = t
                self.trade_history.append(Trade(symbol, action, qty, price, datetime.fromisoformat(ts), com, trade_id, curc))

    def get_portfolio_summary(self):
        self.update_market_data()
        usd_to_gbp = self.get_fx_rate_usd_to_gbp()
        
        total_long_gbp = 0.0
        total_short_gbp = 0.0
        unrealized_pnl_gbp = 0.0
        realized_pnl_gbp = 0.0
        
        for pos in self.positions.values():
            # Convert GBX to GBP, USD to GBP at summary level
            if pos.currency == 'GBX':
                val_gbp = pos.market_value / 100
                unrealized = pos.unrealized_pnl / 100
                realized = pos.realized_pnl / 100
            elif pos.currency == 'USD':
                val_gbp = pos.market_value * usd_to_gbp
                unrealized = pos.unrealized_pnl * usd_to_gbp
                realized = pos.realized_pnl * usd_to_gbp
            else:
                val_gbp = pos.market_value
                unrealized = pos.unrealized_pnl
                realized = pos.realized_pnl
            
            if pos.position_type == 'LONG':
                total_long_gbp += val_gbp
            elif pos.position_type == 'SHORT':
                total_short_gbp += val_gbp
            
            unrealized_pnl_gbp += unrealized
            realized_pnl_gbp += realized

        net_market_value_gbp = total_long_gbp - total_short_gbp
        total_portfolio_value_gbp = self.cash_gbp + net_market_value_gbp
        
        return {
            'cash_gbp': self.cash_gbp,
            'portfolio_value_gbp': total_portfolio_value_gbp,
            'unrealized_pnl_gbp': unrealized_pnl_gbp,
            'realized_pnl_gbp': realized_pnl_gbp,
            'net_market_value_gbp': net_market_value_gbp,
            'positions_count': len(self.positions),
            'last_updated': self.last_update.strftime("%Y-%m-%d %H:%M:%S") if self.last_update else "Never",
            'usd_to_gbp': usd_to_gbp
        }

    def display_portfolio(self):
        summary = self.get_portfolio_summary()
        print("="*80)
        print(f"CASH (GBP): £{summary['cash_gbp']:.2f}")
        print(f"PORTFOLIO VALUE (GBP): £{summary['portfolio_value_gbp']:.2f}")
        print(f"UNREALIZED PnL (GBP): £{summary['unrealized_pnl_gbp']:.2f}")
        print(f"REALIZED PnL (GBP): £{summary['realized_pnl_gbp']:.2f}")
        print(f"POSITIONS COUNT: {summary['positions_count']}")
        print(f"LAST UPDATED: {summary['last_updated']}")
        print("-"*80)
        if not self.positions:
            print("No open positions.")
            print("="*80)
            return

        print(f"{'Symbol':<10} {'Qty':>10} {'Avg Cost':>12} {'Curr Price':>12} {'Mkt Value':>12} {'Unreal PnL':>12} {'Real PnL':>12} {'Type':>6} {'Currency':>8}")
        for pos in self.positions.values():
            if pos.currency == 'GBX':
                avg_cost_disp = pos.avg_cost / 100
                curr_price_disp = pos.current_price / 100
                mkt_val_disp = pos.market_value / 100
                unreal_pnl_disp = pos.unrealized_pnl / 100
                realized_pnl_disp = pos.realized_pnl / 100
                currency_str = 'GBP'
            else:
                avg_cost_disp = pos.avg_cost
                curr_price_disp = pos.current_price
                mkt_val_disp = pos.market_value
                unreal_pnl_disp = pos.unrealized_pnl
                realized_pnl_disp = pos.realized_pnl
                currency_str = pos.currency or 'USD'
            print(f"{pos.symbol:<10} {pos.quantity:>10.4f} {avg_cost_disp:>12.2f} {curr_price_disp:>12.2f} {mkt_val_disp:>12.2f} {unreal_pnl_disp:>12.2f} {realized_pnl_disp:>12.2f} {pos.position_type:>6} {currency_str:>8}")
        print("="*80)

    def display_trade_history(self, symbol: str = None, limit: int = 20):
        trades = self.trade_history[-limit:]
        if symbol:
            symbol_upper = symbol.upper()
            trades = [t for t in trades if t.symbol.upper() == symbol_upper]
        
        print(f"\n{'Trade History for ' + symbol if symbol else 'Trade History':-^100}")
        print(f"{'Symbol':<10} {'Action':<6} {'Qty':>10} {'Price':>12} {'Comm':>12} {'Value':>12} {'Currency':>9} {'Timestamp'}")
        print("-"*100)
        if not trades:
            print("No trades found.")
            return
        for t in trades:
            trade_value = t.quantity * t.price
            if t.currency == 'GBX':
                price_disp = t.price / 100
                commission_disp = t.commission / 100
                value_disp = trade_value / 100
                currency_str = 'GBP'
                print(f"{t.symbol:<10} {t.action:<6} {t.quantity:>10.4f} £{price_disp:>11.2f} £{commission_disp:>11.2f} £{value_disp:>11.2f} {currency_str:>9} {t.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
            else:
                currency_str = t.currency or 'USD'
                print(f"{t.symbol:<10} {t.action:<6} {t.quantity:>10.4f} ${t.price:>11.2f} ${t.commission:>11.2f} ${trade_value:>11.2f} {currency_str:>9} {t.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")
        print("-"*100)

    def export_to_excel(self):
        try:
            wb = load_workbook(self.excel_path)

            if 'Trades' in wb.sheetnames:
                wb.remove(wb['Trades'])
            trades_ws = wb.create_sheet('Trades', 0)
            trades_ws.append(['Trade ID', 'Symbol', 'Action', 'Quantity', 'Price', 'Commission', 'Value', 'Currency', 'Timestamp'])
            for t in self.trade_history:
                trade_value = t.price * t.quantity
                if t.currency == 'GBX':
                    price_disp = t.price / 100
                    commission_disp = t.commission / 100
                    value_disp = trade_value / 100
                    currency_str = 'GBP'
                else:
                    price_disp = t.price
                    commission_disp = t.commission
                    value_disp = trade_value
                    currency_str = t.currency or 'USD'
                trades_ws.append([
                    t.trade_id, t.symbol, t.action, t.quantity, price_disp, commission_disp, value_disp, currency_str, t.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                ])

            if 'Portfolio' in wb.sheetnames:
                wb.remove(wb['Portfolio'])
            portfolio_ws = wb.create_sheet('Portfolio')
            portfolio_ws.append(['Symbol', 'Quantity', 'Avg Cost', 'Current Price', 'Market Value',
                                 'Unrealized PnL', 'Realized PnL', 'Position Type', 'Currency'])
            self.update_market_data()
            for p in self.positions.values():
                if p.currency == 'GBX':
                    portfolio_ws.append([
                        p.symbol, p.quantity,
                        p.avg_cost / 100,
                        p.current_price / 100,
                        p.market_value / 100,
                        p.unrealized_pnl / 100,
                        p.realized_pnl / 100,
                        p.position_type, 'GBP'
                    ])
                else:
                    portfolio_ws.append([
                        p.symbol, p.quantity,
                        p.avg_cost,
                        p.current_price,
                        p.market_value,
                        p.unrealized_pnl,
                        p.realized_pnl,
                        p.position_type,
                        p.currency or 'USD'
                    ])

            if 'Summary' in wb.sheetnames:
                wb.remove(wb['Summary'])
            summary_ws = wb.create_sheet('Summary')
            summary_ws.append(['Metric', 'Value', 'Currency', 'Last Updated'])
            summary = self.get_portfolio_summary()
            summary_ws.append(['Cash', summary['cash_gbp'], 'GBP', summary['last_updated']])
            summary_ws.append(['Portfolio Value', summary['portfolio_value_gbp'], 'GBP', ''])
            summary_ws.append(['Unrealized PnL', summary['unrealized_pnl_gbp'], 'GBP', ''])
            summary_ws.append(['Realized PnL', summary['realized_pnl_gbp'], 'GBP', ''])
            summary_ws.append(['Positions Count', summary['positions_count'], '', ''])
            
            # Autofit columns
            for ws in [trades_ws, portfolio_ws, summary_ws]:
                self._style_header(ws)
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    adjusted_width = min(max_len + 2, 30)
                    ws.column_dimensions[col[0].column_letter].width = adjusted_width

            wb.save(self.excel_path)
            logging.info(f"Exported data to Excel at {self.excel_path}")
        except Exception as e:
            logging.error(f"Excel export failed: {e}")

# Example usage:
if __name__ == "__main__":
    trader = PaperTradingSystem(initial_cash_gbp=6000)
    print("Paper Trading System initialized.")
    # Example trades:
    # trader.place_order('AAPL', 'BUY', 5)
    # trader.place_order('LLOY.L', 'BUY', 100)
    # trader.display_portfolio()
    # trader.display_trade_history()